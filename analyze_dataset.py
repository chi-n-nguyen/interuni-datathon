import pandas as pd
import openpyxl

# Read the Excel file
file_path = "2025 Allianz Datathon Dataset.xlsx"

# First, let's see what sheets are available
wb = openpyxl.load_workbook(file_path)
print("Available sheets:")
for sheet in wb.sheetnames:
    print(f"- {sheet}")

print("\n" + "="*50)

# Read each sheet and display basic information
for sheet_name in wb.sheetnames:
    print(f"\nSheet: {sheet_name}")
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    print(f"Shape: {df.shape}")
    print(f"Columns: {list(df.columns)}")
    print("\nFirst few rows:")
    print(df.head())
    print("\n" + "-"*30)